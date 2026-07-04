#!/usr/bin/env python3
"""Generate supabase/seed-yalaa-doctor-cases.sql from Dr. Yalaa Excel file."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

XLSX = (
    Path.home()
    / "Downloads"
    / "Telegram Desktop"
    / "د. يعلى سعدي روؤف"
    / "د. يعلى سعدي روؤف"
    / "Dr. Yalaa Data .xlsx"
)
OUT = Path(__file__).resolve().parents[1] / "supabase" / "seed-yalaa-doctor-cases.sql"
DOCTOR_USERNAME = "Yalaa.Saadi.Raouf"

FEMALE_FIRST_NAMES = {
    "ابتسام", "اسراء", "اسماء", "امنه", "انتصار", "ايمان", "إيمان", "أماني",
    "امل", "بتول", "بثينة", "ثريا", "جميلة", "جميله", "جنى", "جنان", "حسناء",
    "حنان", "حياة", "خديجة", "دعاء", "رجاء", "رشا", "ريم", "زهراء", "زهره",
    "سجى", "سحر", "سعاد", "سلمى", "سناء", "سهام", "سهى", "سوسن", "سوزان",
    "شذى", "شهد", "شيماء", "صفاء", "عائشة", "عذراء", "غفران", "فاطمة", "فاطمه",
    "فخريه", "كريمه", "كوثر", "ليلى", "لمى", "ماجده", "مروة", "مريم", "مها",
    "منى", "ميساء", "نجاح", "نجلاء", "ندى", "نور", "نوره", "نوال", "هبه", "هبة",
    "هناء", "وفاء", "وداد", "ياسمين", "علياء", "رغد", "رنا", "سميره", "سميه",
    "الاء", "الهام",
}


def sql_str(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def infer_gender(first_name: str) -> str:
    name = first_name.strip()
    if name in FEMALE_FIRST_NAMES:
        return "female"
    if name.endswith("ة") or name.endswith("اء") or name.endswith("ى"):
        return "female"
    return "male"


def load_rows() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        n1 = cell_str(ws.cell(r, 2).value)
        if not n1:
            continue
        n2 = cell_str(ws.cell(r, 3).value)
        n3 = cell_str(ws.cell(r, 4).value)
        n4 = cell_str(ws.cell(r, 5).value)
        age_raw = ws.cell(r, 6).value
        diagnosis = cell_str(ws.cell(r, 7).value) or "Not specified"
        try:
            age_value = int(float(age_raw))
        except (TypeError, ValueError):
            continue
        if age_value <= 0:
            continue
        full_name = " ".join(p for p in [n1, n2, n3, n4] if p)
        rows.append(
            {
                "patient_name1": n1,
                "patient_name2": n2,
                "patient_name3": n3,
                "patient_name4": n4,
                "patient_full_name": full_name,
                "age_value": age_value,
                "gender": infer_gender(n1),
                "diagnosis": diagnosis,
            }
        )
    return rows


def main() -> None:
    if not XLSX.exists():
        print(f"Excel not found: {XLSX}", file=sys.stderr)
        raise SystemExit(1)

    rows = load_rows()
    lines: list[str] = [
        "-- =============================================================================",
        "-- URGEN – Import Dr. Yalaa Saadi Raouf patient cases from Excel",
        "-- Run once in Supabase SQL Editor.",
        "--",
        "-- Doctor portal account (already created):",
        f"--   display_name: Yalaa Saadi Raouf",
        f"--   doctor_username: {DOCTOR_USERNAME}",
        "--",
        f"-- Rows imported: {len(rows)}",
        "--   age_unit: years",
        "--   status: sent (awaiting review — editable by doctor before accept)",
        "--   Tests: NOT imported — add later via doctor portal while status = sent",
        "-- =============================================================================",
        "",
        "do $$",
        "declare",
        "  v_doctor_id uuid;",
        "  v_count int;",
        "begin",
        f"  select user_id into v_doctor_id",
        f"  from public.doctor_users",
        f"  where doctor_username = {sql_str(DOCTOR_USERNAME)}",
        f"  limit 1;",
        "",
        "  if v_doctor_id is null then",
        f"    raise exception 'Doctor not found: doctor_username = {DOCTOR_USERNAME}';",
        "  end if;",
        "",
        "  insert into public.doctor_cases (",
        "    doctor_user_id,",
        "    patient_name1, patient_name2, patient_name3, patient_name4,",
        "    patient_full_name,",
        "    age_value, age_unit, gender,",
        "    diagnosis, disease_type,",
        "    oncology_tumor_type, oncology_stage, oncology_treatment,",
        "    status",
        "  ) values",
    ]

    value_rows = []
    for row in rows:
        value_rows.append(
            "  (\n"
            f"    v_doctor_id,\n"
            f"    {sql_str(row['patient_name1'])}, {sql_str(row['patient_name2'])}, "
            f"{sql_str(row['patient_name3'])}, {sql_str(row['patient_name4'])},\n"
            f"    {sql_str(row['patient_full_name'])},\n"
            f"    {row['age_value']}, 'years', {sql_str(row['gender'])},\n"
            f"    {sql_str(row['diagnosis'])}, 'oncology',\n"
            f"    {sql_str(row['diagnosis'])}, 'Not specified', 'Not specified',\n"
            f"    'sent'\n"
            "  )"
        )

    lines.append(",\n".join(value_rows) + ";")
    lines.extend(
        [
            "",
            "  get diagnostics v_count = row_count;",
            "  raise notice 'Imported % doctor_cases for Dr. Yalaa (user_id=%)', v_count, v_doctor_id;",
            "end $$;",
            "",
        ]
    )

    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {len(rows)} rows -> {OUT}")


if __name__ == "__main__":
    main()
