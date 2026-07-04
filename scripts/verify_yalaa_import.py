#!/usr/bin/env python3
"""Compare Excel patient rows vs seed-yalaa-doctor-cases.sql."""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SQL = ROOT / "supabase" / "seed-yalaa-doctor-cases.sql"


def find_xlsx() -> Path:
    base = Path.home() / "Downloads" / "Telegram Desktop"
    matches = list(base.rglob("Dr. Yalaa Data .xlsx"))
    if not matches:
        raise FileNotFoundError("Dr. Yalaa Data .xlsx not found")
    return matches[0]


def cell(value) -> str:
    return "" if value is None else str(value).strip()


def full_name(n1: str, n2: str, n3: str, n4: str) -> str:
    return " ".join(p for p in [n1, n2, n3, n4] if p)


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8")
    xlsx = find_xlsx()
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active

    all_data_rows: list[tuple] = []
    imported: list[tuple] = []
    skipped: list[tuple] = []

    for r in range(2, ws.max_row + 1):
        serial = ws.cell(r, 1).value
        n1 = cell(ws.cell(r, 2).value)
        n2 = cell(ws.cell(r, 3).value)
        n3 = cell(ws.cell(r, 4).value)
        n4 = cell(ws.cell(r, 5).value)
        age = ws.cell(r, 6).value
        diag = cell(ws.cell(r, 7).value)
        test = cell(ws.cell(r, 8).value)

        if not any([n1, n2, n3, n4, age, diag, test]):
            continue

        row = (r, serial, n1, n2, n3, n4, age, diag, test)
        all_data_rows.append(row)

        if not n1:
            skipped.append((*row, "no first name"))
            continue
        try:
            age_i = int(float(age))
            if age_i <= 0:
                skipped.append((*row, "invalid age"))
                continue
        except (TypeError, ValueError):
            skipped.append((*row, "missing/invalid age"))
            continue

        imported.append(
            (
                r,
                serial,
                n1,
                n2,
                n3,
                n4,
                age_i,
                diag or "Not specified",
                full_name(n1, n2, n3, n4),
            )
        )

    sql_text = SQL.read_text(encoding="utf-8")
    sql_rows = re.findall(
        r"patient_full_name,\s*\n\s*(\d+), 'years'",
        sql_text,
    )
    sql_names = re.findall(r"patient_full_name,\s*\n\s*'([^']*)'", sql_text)
    # Better: count value tuples ending with 'sent'
    sql_count = len(re.findall(r"\n    'sent'\n", sql_text))

    print(f"Excel file: {xlsx}")
    print(f"Sheet max_row: {ws.max_row}")
    print(f"Rows with data (non-empty): {len(all_data_rows)}")
    print(f"Imported by script rules: {len(imported)}")
    print(f"Skipped by script rules: {len(skipped)}")
    print(f"SQL seed file rows: {sql_count}")
    print()

    serials = [x[1] for x in all_data_rows if x[1] is not None]
    if serials:
        print(f"Serial numbers in sheet: min={min(serials)}, max={max(serials)}")

    if skipped:
        print("\n--- Skipped rows ---")
        for row in skipped:
            r, serial, n1, n2, n3, n4, age, diag, test, reason = row
            print(
                f"  row {r} | serial={serial} | {full_name(n1,n2,n3,n4)} | age={age} | {reason}"
            )

    # Check each imported name exists in SQL
    missing_in_sql: list[tuple] = []
    for row in imported:
        r, serial, n1, n2, n3, n4, age_i, diag, fname = row
        pattern = re.escape(fname)
        if not re.search(rf"{pattern}", sql_text):
            missing_in_sql.append(row)

    print(f"\nImported rows missing from SQL text: {len(missing_in_sql)}")
    if missing_in_sql:
        for row in missing_in_sql[:20]:
            print(f"  serial={row[1]} | {row[8]} | age={row[6]}")

    # Duplicate full names in excel
    from collections import Counter

    name_counts = Counter(x[8] for x in imported)
    dups = [(n, c) for n, c in name_counts.items() if c > 1]
    if dups:
        print(f"\nDuplicate full names in Excel ({len(dups)}):")
        for n, c in sorted(dups, key=lambda x: -x[1])[:15]:
            print(f"  {c}x {n}")

    if len(imported) == sql_count:
        print("\nRESULT: SQL matches all importable Excel rows.")
    else:
        print(f"\nRESULT: MISMATCH — Excel importable={len(imported)}, SQL={sql_count}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
